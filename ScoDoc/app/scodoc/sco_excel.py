# -*- mode: python -*-
# -*- coding: utf-8 -*-

##############################################################################
#
# Gestion scolarite IUT
#
# Copyright (c) 1999 - 2021 Emmanuel Viennet.  All rights reserved.
#
# This program is free software; you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation; either version 2 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program; if not, write to the Free Software
# Foundation, Inc., 59 Temple Place, Suite 330, Boston, MA  02111-1307  USA
#
#   Emmanuel Viennet      emmanuel.viennet@viennet.net
#
##############################################################################


""" Excel file handling
"""
import datetime
import io
import time
from enum import Enum
from tempfile import NamedTemporaryFile

import openpyxl.utils.datetime
from openpyxl.styles.numbers import FORMAT_NUMBER_00, FORMAT_GENERAL, FORMAT_DATE_DDMMYY
from openpyxl.comments import Comment
from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

import app.scodoc.sco_utils as scu
from app.scodoc import notesdb
from app.scodoc import sco_preferences
from app import log
from app.scodoc.sco_exceptions import ScoValueError


class COLORS(Enum):
    BLACK = "FF000000"
    WHITE = "FFFFFFFF"
    RED = "FFFF0000"
    BROWN = "FF993300"
    PURPLE = "FF993366"
    BLUE = "FF0000FF"
    ORANGE = "FFFF3300"
    LIGHT_YELLOW = "FFFFFF99"


# Un style est enregistré comme un dictionnaire qui précise la valeur d'un attributdans la liste suivante:
# font, border, number_format, fill,...
# (cf https://openpyxl.readthedocs.io/en/stable/styles.html#working-with-styles)


def xldate_as_datetime(xldate, datemode=0):
    """Conversion d'une date Excel en datetime python
    Deux formats de chaîne acceptés:
     * JJ/MM/YYYY (chaîne naïve)
     * Date ISO (valeur de type date lue dans la feuille)
    Peut lever une ValueError
    """
    try:
        return datetime.datetime.strptime(xldate, "%d/%m/%Y")
    except:
        return openpyxl.utils.datetime.from_ISO8601(xldate)


def adjust_sheetname(sheet_name):
    """Renvoie un nom convenable pour une feuille excel: < 31 cars, sans caractères spéciaux
    Le / n'est pas autorisé par exemple.
    Voir https://xlsxwriter.readthedocs.io/workbook.html#add_worksheet
    """
    sheet_name = scu.make_filename(sheet_name)
    # Le nom de la feuille ne peut faire plus de 31 caractères.
    # si la taille du nom de feuille est > 31 on tronque (on pourrait remplacer par 'feuille' ?)
    return sheet_name[:31]


class ScoExcelBook:
    """Permet la génération d'un classeur xlsx composé de plusieurs feuilles.
    usage:
        wb = ScoExcelBook()
        ws0 = wb.create_sheet('sheet name 0')
        ws1 = wb.create_sheet('sheet name 1')
        ...
        steam = wb.generate()
    """

    def __init__(self):
        self.sheets = []  # list of sheets
        self.wb = Workbook(write_only=True)

    def create_sheet(self, sheet_name="feuille", default_style=None):
        """Crée une nouvelle feuille dans ce classeur
        sheet_name -- le nom de la feuille
        default_style -- le style par défaut
        """
        sheet_name = adjust_sheetname(sheet_name)
        ws = self.wb.create_sheet(sheet_name)
        sheet = ScoExcelSheet(sheet_name, default_style, ws)
        self.sheets.append(sheet)
        return sheet

    def generate(self):
        """génération d'un stream binaire représentant la totalité du classeur.
        retourne le flux
        """
        for sheet in self.sheets:
            sheet.prepare()
        # construction d'un flux
        # (https://openpyxl.readthedocs.io/en/stable/tutorial.html#saving-as-a-stream)
        with NamedTemporaryFile() as tmp:
            self.wb.save(tmp.name)
            tmp.seek(0)
            return tmp.read()


def excel_make_style(
    bold=False,
    italic=False,
    outline=False,
    color: COLORS = COLORS.BLACK,
    bgcolor: COLORS = None,
    halign=None,
    valign=None,
    number_format=None,
    font_name="Arial",
    size=10,
):
    """Contruit un style.
    Les couleurs peuvent être spécfiées soit par une valeur de COLORS,
    soit par une chaine argb (exple "FF00FF00" pour le vert)
    color -- La couleur du texte
    bgcolor -- la couleur de fond
    halign -- alignement horizontal ("left", "right", "center")
    valign -- alignement vertical ("top", "bottom", "center")
    number_format -- formattage du contenu ("General", "@", ...)
    font_name -- police
    size -- taille de police
    """
    style = {}
    font = Font(
        name=font_name,
        bold=bold,
        italic=italic,
        outline=outline,
        color=color.value,
        size=size,
    )
    style["font"] = font
    if bgcolor:
        style["fill"] = PatternFill(fill_type="solid", fgColor=bgcolor.value)
    if halign or valign:
        al = Alignment()
        if halign:
            al.horizontal = {
                "left": "left",
                "right": "right",
                "center": "center",
            }[halign]
        if valign:
            al.vertical = {
                "top": "top",
                "bottom": "bottom",
                "center": "center",
            }[valign]
        style["alignment"] = al
    if number_format is None:
        style["number_format"] = FORMAT_GENERAL
    else:
        style["number_format"] = number_format
    return style


class ScoExcelSheet:
    """Représente une feuille qui peut être indépendante ou intégrée dans un SCoExcelBook.
    En application des directives de la bibliothèque sur l'écriture optimisée, l'ordre des opérations
    est imposé:
    * instructions globales (largeur/maquage des colonnes et ligne, ...)
    * construction et ajout des cellules et ligne selon le sens de lecture (occidental)
    ligne de haut en bas et cellules de gauche à droite (i.e. A1, A2, .. B1, B2, ..)
    * pour finit appel de la méthode de génération
    """

    def __init__(self, sheet_name="feuille", default_style=None, wb=None):
        """Création de la feuille. sheet_name
        -- le nom de la feuille default_style
        -- le style par défaut des cellules ws
        -- None si la feuille est autonome (dans ce cas ell crée son propre wb), sinon c'est la worksheet
        créée par le workbook propriétaire un workbook est crée et associé à cette feuille.
        """
        # Le nom de la feuille ne peut faire plus de 31 caractères.
        # si la taille du nom de feuille est > 31 on tronque (on pourrait remplacer par 'feuille' ?)
        self.sheet_name = adjust_sheetname(sheet_name)
        if default_style is None:
            default_style = excel_make_style()
        self.default_style = default_style
        if wb is None:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = self.sheet_name
        else:
            self.wb = None
            self.ws = wb
        # internal data
        self.rows = []  # list of list of cells
        self.column_dimensions = {}
        self.row_dimensions = {}

    def excel_make_composite_style(
        self,
        alignment=None,
        border=None,
        fill=None,
        number_format=None,
        font=None,
    ):
        style = {}
        if font is not None:
            style["font"] = font
        if alignment is not None:
            style["alignment"] = alignment
        if border is not None:
            style["border"] = border
        if fill is not None:
            style["fill"] = fill
        if number_format is None:
            style["number_format"] = FORMAT_GENERAL
        else:
            style["number_format"] = number_format
        return style

    @staticmethod
    def i2col(idx):
        if idx < 26:  # one letter key
            return chr(idx + 65)
        else:  # two letters AA..ZZ
            first = (idx // 26) + 66
            second = (idx % 26) + 65
            return "" + chr(first) + chr(second)

    def set_column_dimension_width(self, cle=None, value=21):
        """Détermine la largeur d'une colonne. cle -- identifie la colonne ("A" "B", ... ou 0, 1, 2, ...) si None,
        value donne la liste des largeurs de colonnes depuis A, B, C, ... value -- la dimension (unité : 7 pixels
        comme affiché dans Excel)
        """
        if cle is None:
            for i, val in enumerate(value):
                self.ws.column_dimensions[self.i2col(i)].width = val
            # No keys: value is a list of widths
        elif type(cle) == str:  # accepts set_column_with("D", ...)
            self.ws.column_dimensions[cle].width = value
        else:
            self.ws.column_dimensions[self.i2col(cle)].width = value

    def set_row_dimension_height(self, cle=None, value=21):
        """Détermine la hauteur d'une ligne. cle -- identifie la ligne (1, 2, ...) si None,
        value donne la liste des hauteurs de colonnes depuis 1, 2, 3, ... value -- la dimension
        """
        if cle is None:
            for i, val in enumerate(value, start=1):
                self.ws.row_dimensions[i].height = val
            # No keys: value is a list of widths
        else:
            self.ws.row_dimensions[cle].height = value

    def set_row_dimension_hidden(self, cle, value):
        """Masque ou affiche une ligne.
        cle -- identifie la colonne (1...)
        value -- boolean (vrai = colonne cachée)
        """
        self.ws.row_dimensions[cle].hidden = value

    def make_cell(self, value: any = None, style=None, comment=None):
        """Construit une cellule.
        value -- contenu de la cellule (texte, numérique, booléen ou date)
        style -- style par défaut (dictionnaire cf. excel_make_style) de la feuille si non spécifié
        """
        # adapatation des valeurs si nécessaire
        if value is None:
            value = ""
        elif value is True:
            value = 1
        elif value is False:
            value = 0
        elif isinstance(value, datetime.datetime):
            value = value.replace(
                tzinfo=None
            )  # make date naive (cf https://openpyxl.readthedocs.io/en/latest/datetime.html#timezones)

        # création de la cellule
        cell = WriteOnlyCell(self.ws, value)

        # recopie des styles
        if style is None:
            style = self.default_style
        if "font" in style:
            cell.font = style["font"]
        if "alignment" in style:
            cell.alignment = style["alignment"]
        if "border" in style:
            cell.border = style["border"]
        if "fill" in style:
            cell.fill = style["fill"]
        if "number_format" in style:
            cell.number_format = style["number_format"]
        if "fill" in style:
            cell.fill = style["fill"]
        if "alignment" in style:
            cell.alignment = style["alignment"]
        if not comment is None:
            cell.comment = Comment(comment, "scodoc")
            lines = comment.splitlines()
            cell.comment.width = 7 * max([len(line) for line in lines])
            cell.comment.height = 20 * len(lines)

        # test datatype to overwrite datetime format
        if isinstance(value, datetime.date):
            cell.data_type = "d"
            cell.number_format = FORMAT_DATE_DDMMYY
        elif isinstance(value, int) or isinstance(value, float):
            cell.data_type = "n"
        else:
            cell.data_type = "s"

        return cell

    def make_row(self, values: list, style=None, comments=None):
        # TODO make possible differents styles in a row
        if comments is None:
            comments = [None] * len(values)
        return [
            self.make_cell(value, style, comment)
            for value, comment in zip(values, comments)
        ]

    def append_single_cell_row(self, value: any, style=None):
        """construit une ligne composée d'une seule cellule et l'ajoute à la feuille.
        mêmes paramètres que make_cell:
        value -- contenu de la cellule (texte ou numérique)
        style -- style par défaut de la feuille si non spécifié
        """
        self.append_row([self.make_cell(value, style)])

    def append_blank_row(self):
        """construit une ligne vide et l'ajoute à la feuille."""
        self.append_row([None])

    def append_row(self, row):
        """ajoute une ligne déjà construite à la feuille."""
        self.rows.append(row)

    def prepare(self):
        """génére un flux décrivant la feuille.
        Ce flux pourra ensuite être repris dans send_excel_file (classeur mono feille)
        ou pour la génération d'un classeur multi-feuilles
        """
        for row in self.column_dimensions.keys():
            self.ws.column_dimensions[row] = self.column_dimensions[row]
        for row in self.row_dimensions.keys():
            self.ws.row_dimensions[row] = self.row_dimensions[row]
        for row in self.rows:
            self.ws.append(row)

    def generate(self):
        """génération d'un classeur mono-feuille"""
        # this method makes sense only if it is a standalone worksheet (else call workbook.generate()
        if self.wb is None:  # embeded sheet
            raise ScoValueError("can't generate a single sheet from a ScoWorkbook")

        # construction d'un flux (https://openpyxl.readthedocs.io/en/stable/tutorial.html#saving-as-a-stream)
        self.prepare()
        with NamedTemporaryFile() as tmp:
            self.wb.save(tmp.name)
            tmp.seek(0)
            return tmp.read()


def excel_simple_table(
    titles=None, lines=None, sheet_name=b"feuille", titles_styles=None, comments=None
):
    """Export simple type 'CSV': 1ere ligne en gras, le reste tel quel"""
    ws = ScoExcelSheet(sheet_name)
    if titles is None:
        titles = []
    if lines is None:
        lines = [[]]
    if titles_styles is None:
        style = excel_make_style(bold=True)
        titles_styles = [style] * len(titles)
    if comments is None:
        comments = [None] * len(titles)
    # ligne de titres
    ws.append_row(
        [
            ws.make_cell(it, style, comment)
            for (it, style, comment) in zip(titles, titles_styles, comments)
        ]
    )
    default_style = excel_make_style()
    text_style = excel_make_style(number_format=FORMAT_GENERAL)
    int_style = excel_make_style()
    float_style = excel_make_style(number_format=FORMAT_NUMBER_00)
    for line in lines:
        cells = []
        for it in line:
            cell_style = default_style
            if type(it) == float:
                cell_style = float_style
            elif type(it) == int:
                cell_style = int_style
            else:
                cell_style = text_style
            cells.append(ws.make_cell(it, cell_style))
        ws.append_row(cells)
    return ws.generate()


def excel_feuille_saisie(e, titreannee, description, lines):
    """Genere feuille excel pour saisie des notes.
    E: evaluation (dict)
    lines: liste de tuples
               (etudid, nom, prenom, etat, groupe, val, explanation)
    """
    sheet_name = "Saisie notes"
    ws = ScoExcelSheet(sheet_name)

    # ajuste largeurs colonnes (unite inconnue, empirique)
    ws.set_column_dimension_width("A", 11.0 / 7)  # codes
    # ws.set_column_dimension_hidden("A", True)  # codes
    ws.set_column_dimension_width("B", 164.00 / 7)  # noms
    ws.set_column_dimension_width("C", 109.0 / 7)  # prenoms
    ws.set_column_dimension_width("D", 164.0 / 7)  # groupes
    ws.set_column_dimension_width("E", 115.0 / 7)  # notes
    ws.set_column_dimension_width("F", 355.0 / 7)  # remarques

    # fontes
    font_base = Font(name="Arial", size=12)
    font_bold = Font(name="Arial", bold=True)
    font_italic = Font(name="Arial", size=12, italic=True, color=COLORS.RED.value)
    font_titre = Font(name="Arial", bold=True, size=14)
    font_purple = Font(name="Arial", color=COLORS.PURPLE.value)
    font_brown = Font(name="Arial", color=COLORS.BROWN.value)
    font_blue = Font(name="Arial", size=9, color=COLORS.BLUE.value)

    # bordures
    side_thin = Side(border_style="thin", color=COLORS.BLACK.value)
    border_top = Border(top=side_thin)
    border_right = Border(right=side_thin)

    # fonds
    fill_light_yellow = PatternFill(
        patternType="solid", fgColor=COLORS.LIGHT_YELLOW.value
    )

    # styles
    style = {"font": font_base}
    style_titres = {"font": font_titre}
    style_expl = {"font": font_italic}

    style_ro = {  # cells read-only
        "font": font_purple,
        "border": border_right,
    }
    style_dem = {
        "font": font_brown,
        "border": border_top,
    }
    style_nom = {  # style pour nom, prenom, groupe
        "font": font_base,
        "border": border_top,
    }
    style_notes = {
        "font": font_bold,
        "number_format": FORMAT_GENERAL,
        "fill": fill_light_yellow,
        "border": border_top,
    }
    style_comment = {
        "font": font_blue,
        "border": border_top,
    }

    # ligne de titres
    ws.append_single_cell_row(
        "Feuille saisie note (à enregistrer au format excel)", style_titres
    )
    # lignes d'instructions
    ws.append_single_cell_row(
        "Saisir les notes dans la colonne E (cases jaunes)", style_expl
    )
    ws.append_single_cell_row("Ne pas modifier les cases en mauve !", style_expl)
    # Nom du semestre
    ws.append_single_cell_row(scu.unescape_html(titreannee), style_titres)
    # description evaluation
    ws.append_single_cell_row(scu.unescape_html(description), style_titres)
    ws.append_single_cell_row(
        "Evaluation du %s (coef. %g)" % (e["jour"], e["coefficient"]), style
    )
    # ligne blanche
    ws.append_blank_row()
    # code et titres colonnes
    ws.append_row(
        [
            ws.make_cell("!%s" % e["evaluation_id"], style_ro),
            ws.make_cell("Nom", style_titres),
            ws.make_cell("Prénom", style_titres),
            ws.make_cell("Groupe", style_titres),
            ws.make_cell("Note sur %g" % e["note_max"], style_titres),
            ws.make_cell("Remarque", style_titres),
        ]
    )

    # etudiants
    for line in lines:
        st = style_nom
        if line[3] != "I":
            st = style_dem
            if line[3] == "D":  # demissionnaire
                s = "DEM"
            else:
                s = line[3]  # etat autre
        else:
            s = line[4]  # groupes TD/TP/...
        try:
            val = float(line[5])
        except ValueError:
            val = line[5]
        ws.append_row(
            [
                ws.make_cell("!" + line[0], style_ro),  # code
                ws.make_cell(line[1], st),
                ws.make_cell(line[2], st),
                ws.make_cell(s, st),
                ws.make_cell(val, style_notes),  # note
                ws.make_cell(line[6], style_comment),  # comment
            ]
        )

    # explication en bas
    ws.append_row([None, ws.make_cell("Code notes", style_titres)])
    ws.append_row(
        [
            None,
            ws.make_cell("ABS", style_expl),
            ws.make_cell("absent (0)", style_expl),
        ]
    )
    ws.append_row(
        [
            None,
            ws.make_cell("EXC", style_expl),
            ws.make_cell("pas prise en compte", style_expl),
        ]
    )
    ws.append_row(
        [
            None,
            ws.make_cell("ATT", style_expl),
            ws.make_cell("en attente", style_expl),
        ]
    )
    ws.append_row(
        [
            None,
            ws.make_cell("SUPR", style_expl),
            ws.make_cell("pour supprimer note déjà entrée", style_expl),
        ]
    )
    ws.append_row(
        [
            None,
            ws.make_cell("", style_expl),
            ws.make_cell("cellule vide -> note non modifiée", style_expl),
        ]
    )
    return ws.generate()


def excel_bytes_to_list(bytes_content):
    try:
        filelike = io.BytesIO(bytes_content)
        return _excel_to_list(filelike)
    except:
        raise ScoValueError(
            """Le fichier xlsx attendu n'est pas lisible !
            Peut-être avez-vous fourni un fichier au mauvais format (txt, xls, ..)
            """
        )


def excel_file_to_list(filename):
    try:
        return _excel_to_list(filename)
    except:
        raise ScoValueError(
            """Le fichier xlsx attendu n'est pas lisible !
            Peut-être avez-vous fourni un fichier au mauvais format (txt, xls, ...)
            """
        )


def _excel_to_list(filelike):
    """returns list of list
    convert_to_string is a conversion function applied to all non-string values (ie numbers)
    """
    try:
        wb = load_workbook(filename=filelike, read_only=True, data_only=True)
    except:
        log("Excel_to_list: failure to import document")
        with open("/tmp/last_scodoc_import_failure" + scu.XLSX_SUFFIX, "wb") as f:
            f.write(filelike)
        raise ScoValueError(
            "Fichier illisible: assurez-vous qu'il s'agit bien d'un document Excel !"
        )
    diag = []  # liste de chaines pour former message d'erreur
    # n'utilise que la première feuille
    if len(wb.get_sheet_names()) < 1:
        diag.append("Aucune feuille trouvée dans le classeur !")
        return diag, None
    if len(wb.get_sheet_names()) > 1:
        diag.append("Attention: n'utilise que la première feuille du classeur !")
    # fill matrix
    sheet_name = wb.get_sheet_names()[0]
    ws = wb.get_sheet_by_name(sheet_name)
    sheet_name = sheet_name.encode(scu.SCO_ENCODING, "backslashreplace")
    values = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                values[(cell.row - 1, cell.column - 1)] = str(cell.value)
    if not values:
        diag.append(
            "Aucune valeur trouvée dans la feuille %s !"
            % sheet_name.decode(scu.SCO_ENCODING)
        )
        return diag, None
    indexes = list(values.keys())
    # search numbers of rows and cols
    rows = [x[0] for x in indexes]
    cols = [x[1] for x in indexes]
    nbcols = max(cols) + 1
    nbrows = max(rows) + 1
    m = []
    for _ in range(nbrows):
        m.append([""] * nbcols)

    for row_idx, col_idx in indexes:
        v = values[(row_idx, col_idx)]
        # if isinstance(v, six.text_type):
        #     v = v.encode(scu.SCO_ENCODING, "backslashreplace")
        # elif convert_to_string:
        #     v = convert_to_string(v)
        m[row_idx][col_idx] = v
    diag.append(
        'Feuille "%s", %d lignes' % (sheet_name.decode(scu.SCO_ENCODING), len(m))
    )
    # diag.append(str(M))
    #
    return diag, m


def excel_feuille_listeappel(
    sem,
    groupname,
    lines,
    partitions=None,
    with_codes=False,
    with_paiement=False,
    server_name=None,
):
    """generation feuille appel"""
    if partitions is None:
        partitions = []
    formsemestre_id = sem["formsemestre_id"]
    sheet_name = "Liste " + groupname

    ws = ScoExcelSheet(sheet_name)
    ws.set_column_dimension_width("A", 3)
    ws.set_column_dimension_width("B", 35)
    ws.set_column_dimension_width("C", 12)

    font1 = Font(name="Arial", size=11)
    font1i = Font(name="Arial", size=10, italic=True)
    font1b = Font(name="Arial", size=11, bold=True)

    side_thin = Side(border_style="thin", color=COLORS.BLACK.value)

    border_tbl = Border(top=side_thin, bottom=side_thin, left=side_thin)
    border_tblr = Border(
        top=side_thin, bottom=side_thin, left=side_thin, right=side_thin
    )

    style1i = {
        "font": font1i,
    }

    style1b = {
        "font": font1,
        "border": border_tbl,
    }

    style2 = {
        "font": Font(name="Arial", size=14),
    }

    style2b = {
        "font": font1i,
        "border": border_tblr,
    }

    style2t3 = {
        "border": border_tblr,
    }

    style2t3bold = {
        "font": font1b,
        "border": border_tblr,
    }

    style3 = {
        "font": Font(name="Arial", bold=True, size=14),
    }

    nb_weeks = 4  # nombre de colonnes pour remplir absences

    # ligne 1
    title = "%s %s (%s - %s)" % (
        sco_preferences.get_preference("DeptName", formsemestre_id),
        notesdb.unquote(sem["titre_num"]),
        sem["date_debut"],
        sem["date_fin"],
    )

    ws.append_row([None, ws.make_cell(title, style2)])

    # ligne 2
    ws.append_row([None, ws.make_cell("Discipline :", style2)])

    # ligne 3
    cell_2 = ws.make_cell("Enseignant :", style2)
    cell_6 = ws.make_cell(("Groupe %s" % groupname), style3)
    ws.append_row([None, cell_2, None, None, None, None, cell_6])

    # ligne 4: Avertissement pour ne pas confondre avec listes notes
    cell_2 = ws.make_cell(
        "Ne pas utiliser cette feuille pour saisir les notes !", style1i
    )
    ws.append_row([None, None, cell_2])

    ws.append_blank_row()
    ws.append_blank_row()

    # ligne 7: Entête (contruction dans une liste cells)
    cell_2 = ws.make_cell("Nom", style3)
    cells = [None, cell_2]
    for partition in partitions:
        cells.append(ws.make_cell(partition["partition_name"], style3))
    if with_codes:
        cells.append(ws.make_cell("etudid", style3))
        cells.append(ws.make_cell("code_nip", style3))
        cells.append(ws.make_cell("code_ine", style3))
    for i in range(nb_weeks):
        cells.append(ws.make_cell("", style2b))
    ws.append_row(cells)

    n = 0
    # pour chaque étudiant
    for t in lines:
        n += 1
        nomprenom = (
            t["civilite_str"] + " " + t["nom"] + " " + t["prenom"].lower().capitalize()
        )
        style_nom = style2t3
        if with_paiement:
            paie = t.get("paiementinscription", None)
            if paie is None:
                nomprenom += " (inscription ?)"
                style_nom = style2t3bold
            elif not paie:
                nomprenom += " (non paiement)"
                style_nom = style2t3bold
        cell_1 = ws.make_cell(n, style1b)
        cell_2 = ws.make_cell(nomprenom, style_nom)
        cells = [cell_1, cell_2]

        for partition in partitions:
            if partition["partition_name"]:
                cells.append(
                    ws.make_cell(t.get(partition["partition_id"], ""), style2t3)
                )
        if with_codes:
            cells.append(ws.make_cell(t["etudid"], style2t3))
            code_nip = t.get("code_nip", "")
            cells.append(ws.make_cell(code_nip, style2t3))
            code_ine = t.get("code_ine", "")
            cells.append(ws.make_cell(code_ine, style2t3))
        cells.append(ws.make_cell(t.get("etath", ""), style2b))
        for i in range(1, nb_weeks):
            cells.append(ws.make_cell(style=style2t3))
        ws.append_row(cells)

    ws.append_blank_row()

    # bas de page (date, serveur)
    dt = time.strftime("%d/%m/%Y à %Hh%M")
    if server_name:
        dt += " sur " + server_name
    cell_2 = ws.make_cell(("Liste éditée le " + dt), style1i)
    ws.append_row([None, cell_2])

    return ws.generate()
