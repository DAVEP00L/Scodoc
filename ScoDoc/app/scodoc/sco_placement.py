# -*- mode: python -*-
# -*- coding: utf-8 -*-

##############################################################################
#
# Gestion scolarite IUT
#
# Copyright (c) 1999 - 2021 Emmanuel Viennet.  All rights reserved.
#
# This program is free software; you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation; either version 2 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program; if not, write to the Free Software
# Foundation, Inc., 59 Temple Place, Suite 330, Boston, MA  02111-1307  USA
#
#   Emmanuel Viennet      emmanuel.viennet@viennet.net
#
##############################################################################

"""ScoDoc: génération feuille émargement et placement

Contribution J.-M. Place 2021
basée sur une idée de M. Salomon, UFC / IUT DE BELFORT-MONTBÉLIARD, 2016

"""
import random
import time
from copy import copy

import wtforms.validators
from flask import request, render_template
from flask_login import current_user
from flask_wtf import FlaskForm
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from wtforms import (
    StringField,
    SubmitField,
    SelectField,
    RadioField,
    HiddenField,
    SelectMultipleField,
)
import app.scodoc.sco_utils as scu
import app.scodoc.notesdb as ndb
from app import ScoValueError
from app.scodoc import html_sco_header, sco_preferences
from app.scodoc import sco_edit_module
from app.scodoc import sco_evaluations
from app.scodoc import sco_excel
from app.scodoc.sco_excel import ScoExcelBook, COLORS
from app.scodoc import sco_formsemestre
from app.scodoc import sco_formsemestre_inscriptions
from app.scodoc import sco_groups
from app.scodoc import sco_moduleimpl
from app.scodoc import sco_permissions_check
from app.scodoc.gen_tables import GenTable
from app.scodoc import sco_etud
import sco_version

_ = lambda x: x  # sans babel
_l = _

COORD = "Coordonnées"
SEQ = "Continue"

TOUS = "Tous"


def _get_group_info(evaluation_id):
    # groupes
    groups = sco_groups.do_evaluation_listegroupes(evaluation_id, include_default=True)
    has_groups = False
    groups_tree = {}
    for group in groups:
        partition = group["partition_name"] or TOUS
        group_id = group["group_id"]
        group_name = group["group_name"] or TOUS
        if partition not in groups_tree:
            groups_tree[partition] = {}
        groups_tree[partition][group_name] = group_id
        if partition != TOUS:
            has_groups = True
        else:
            has_groups = False
    nb_groups = sum([len(groups_tree[p]) for p in groups_tree])
    return groups_tree, has_groups, nb_groups


class PlacementForm(FlaskForm):
    """Formulaire pour placement des étudiants en Salle"""

    evaluation_id = HiddenField("evaluation_id")
    file_format = RadioField(
        "Format de fichier",
        choices=["pdf", "xls"],
        validators=[
            wtforms.validators.DataRequired("indiquez le format du fichier attendu"),
        ],
    )
    surveillants = StringField("Surveillants", validators=[])
    batiment = StringField("Batiment")
    salle = StringField("Salle")
    nb_rangs = SelectField(
        "nb de places en largeur",
        coerce=int,
        choices=[3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14],
        description="largeur de la salle, en nombre de places",
    )
    etiquetage = RadioField(
        "Numérotation",
        choices=[SEQ, COORD],
        validators=[
            wtforms.validators.DataRequired("indiquez le style de  numérotation"),
        ],
    )
    groups = SelectMultipleField(
        "Groupe(s)",
        validators=[],
    )
    submit = SubmitField("OK")

    def __init__(self, formdata=None, data=None):
        super().__init__(formdata=formdata, data=data)
        self.groups_tree = {}
        self.has_groups = None
        self.nb_groups = None
        self.tous_id = None
        self.set_evaluation_infos(data["evaluation_id"])

    def set_evaluation_infos(self, evaluation_id):
        """Initialise les données du formulaire avec les données de l'évaluation."""
        eval_data = sco_evaluations.do_evaluation_list({"evaluation_id": evaluation_id})
        if not eval_data:
            raise ScoValueError("invalid evaluation_id")
        self.groups_tree, self.has_groups, self.nb_groups = _get_group_info(
            evaluation_id
        )
        choices = []
        for partition in self.groups_tree:
            for groupe in self.groups_tree[partition]:
                if (
                    groupe == TOUS
                ):  #  Affichage et valeur spécifique pour le groupe TOUS
                    self.tous_id = str(self.groups_tree[partition][groupe])
                    choices.append((TOUS, TOUS))
                else:
                    groupe_id = str(self.groups_tree[partition][groupe])
                    choices.append((groupe_id, "%s (%s)" % (str(groupe), partition)))
        self.groups.choices = choices
        # self.groups.default = [TOUS]  # Ne fonctionnne pas... (ni dans la déclaration de PlaceForm.groups)
        # la réponse [] est de toute façon transposée en [ self.tous_id ] lors du traitement (cas du groupe unique)


class _DistributeurContinu:
    """Distribue les places selon un ordre numérique."""

    def __init__(self):
        self.position = 1

    def suivant(self):
        """Retounre la désignation de la place suivante"""
        retour = self.position
        self.position += 1
        return retour


class _Distributeur2D:
    """Distribue les places selon des coordonnées sur nb_rangs."""

    def __init__(self, nb_rangs):
        self.nb_rangs = nb_rangs
        self.rang = 1
        self.index = 1

    def suivant(self):
        """Retounre la désignation de la place suivante"""
        retour = (self.index, self.rang)
        self.rang += 1
        if self.rang > self.nb_rangs:
            self.rang = 1
            self.index += 1
        return retour


def placement_eval_selectetuds(evaluation_id):
    """Creation de l'écran de placement"""
    form = PlacementForm(
        request.form,
        data={"evaluation_id": int(evaluation_id), "groups": TOUS},
    )
    if form.validate_on_submit():
        runner = PlacementRunner(form)
        if not runner.check_placement():
            return (
                """<h2>Génération du placement impossible pour %s</h2>
            <p>(vérifiez que le semestre n'est pas verrouillé et que vous
            avez l'autorisation d'effectuer cette opération)</p>
            <p><a href="moduleimpl_status?moduleimpl_id=%s">Continuer</a></p>
            """
                % runner.__dict__
            )
        return runner.exec_placement()  # calcul et generation du fichier
    htmls = [
        html_sco_header.sco_header(init_jquery_ui=True),
        sco_evaluations.evaluation_describe(evaluation_id=evaluation_id),
        "<h3>Placement et émargement des étudiants</h3>",
        render_template("scodoc/forms/placement.html", form=form),
    ]
    footer = html_sco_header.sco_footer()
    return "\n".join(htmls) + "<p>" + footer


class PlacementRunner:
    """Execution de l'action définie par le formulaire"""

    def __init__(self, form):
        """Calcul et génération du fichier sur la base des données du formulaire"""
        self.evaluation_id = form["evaluation_id"].data
        self.etiquetage = form["etiquetage"].data
        self.surveillants = form["surveillants"].data
        self.batiment = form["batiment"].data
        self.salle = form["salle"].data
        self.nb_rangs = form["nb_rangs"].data
        self.file_format = form["file_format"].data
        if len(form["groups"].data) == 0:
            self.groups_ids = [form.tous_id]
        else:  # On remplace le mot-clé TOUS le l'identiant de ce groupe
            self.groups_ids = [
                gid if gid != TOUS else form.tous_id for gid in form["groups"].data
            ]
        self.eval_data = sco_evaluations.do_evaluation_list(
            {"evaluation_id": self.evaluation_id}
        )[0]
        self.groups = sco_groups.listgroups(self.groups_ids)
        self.gr_title_filename = sco_groups.listgroups_filename(self.groups)
        # gr_title = sco_groups.listgroups_abbrev(d['groups'])
        self.current_user = current_user
        self.moduleimpl_id = self.eval_data["moduleimpl_id"]
        self.moduleimpl_data = sco_moduleimpl.moduleimpl_list(
            moduleimpl_id=self.moduleimpl_id
        )[0]
        self.module_data = sco_edit_module.module_list(
            args={"module_id": self.moduleimpl_data["module_id"]}
        )[0]
        self.sem = sco_formsemestre.get_formsemestre(
            self.moduleimpl_data["formsemestre_id"]
        )
        self.evalname = "%s-%s" % (
            self.module_data["code"],
            ndb.DateDMYtoISO(self.eval_data["jour"]),
        )
        if self.eval_data["description"]:
            self.evaltitre = self.eval_data["description"]
        else:
            self.evaltitre = "évaluation du %s" % self.eval_data["jour"]
        self.desceval = [  # une liste de chaines: description de l'evaluation
            "%s" % self.sem["titreannee"],
            "Module : %s - %s" % (self.module_data["code"], self.module_data["abbrev"]),
            "Surveillants : %s" % self.surveillants,
            "Batiment : %(batiment)s - Salle : %(salle)s" % self.__dict__,
            "Controle : %s (coef. %g)"
            % (self.evaltitre, self.eval_data["coefficient"]),
        ]
        self.styles = None
        self.plan = None
        self.listetud = None

    def check_placement(self):
        """Vérifie que l'utilisateur courant a le droit d'édition sur les notes"""
        # Check access (admin, respformation, and responsable_id)
        return sco_permissions_check.can_edit_notes(
            self.current_user, self.moduleimpl_id
        )

    def exec_placement(self):
        """Excéute l'action liée au formulaire"""
        self._repartition()
        if self.file_format == "xls":
            return self._production_xls()
        return self._production_pdf()

    def _repartition(self):
        """
        Calcule le placement. retourne une liste de couples ((nom, prenom), position)
        """
        # Construit liste des etudiants et les réparti
        self.groups = sco_groups.listgroups(self.groups_ids)
        self.listetud = self._build_listetud()
        self.plan = self._affectation_places()

    def _build_listetud(self):
        get_all_students = None in [
            g["group_name"] for g in self.groups
        ]  # tous les etudiants
        etudids = sco_groups.do_evaluation_listeetuds_groups(
            self.evaluation_id,
            self.groups,
            getallstudents=get_all_students,
            include_dems=True,
        )
        listetud = []  # liste de couples (nom,prenom)
        for etudid in etudids:
            # infos identite etudiant (xxx sous-optimal: 1/select par etudiant)
            ident = sco_etud.etudident_list(ndb.GetDBConnexion(), {"etudid": etudid})[0]
            # infos inscription
            inscr = sco_formsemestre_inscriptions.do_formsemestre_inscription_list(
                {
                    "etudid": etudid,
                    "formsemestre_id": self.moduleimpl_data["formsemestre_id"],
                }
            )[0]
            if inscr["etat"] != "D":
                nom = ident["nom"].upper()
                prenom = ident["prenom"].lower().capitalize()
                etudid = ident["etudid"]
                listetud.append((nom, prenom, etudid))
        random.shuffle(listetud)
        return listetud

    def _affectation_places(self):
        plan = []
        if self.etiquetage == SEQ:
            distributeur = _DistributeurContinu()
        else:
            distributeur = _Distributeur2D(self.nb_rangs)
        for etud in self.listetud:
            plan.append((etud, distributeur.suivant()))
        return plan

    def _production_xls(self):
        filename = "placement_%s_%s" % (self.evalname, self.gr_title_filename)
        xls = self._excel_feuille_placement()
        return scu.send_file(xls, filename, scu.XLSX_SUFFIX, mime=scu.XLSX_MIMETYPE)

    def _production_pdf(self):
        pdf_title = "<br/>".join(self.desceval)
        pdf_title += (
            "\nDate : %(jour)s - Horaire : %(heure_debut)s à %(heure_fin)s"
            % self.eval_data
        )
        filename = "placement_%(evalname)s_%(gr_title_filename)s" % self.__dict__
        titles = {
            "nom": "Nom",
            "prenom": "Prenom",
            "colonne": "Colonne",
            "ligne": "Ligne",
            "place": "Place",
        }
        if self.etiquetage == COORD:
            columns_ids = ["nom", "prenom", "colonne", "ligne"]
        else:
            columns_ids = ["nom", "prenom", "place"]

        rows = []
        for etud in sorted(self.plan, key=lambda item: item[0][0]):  # sort by name
            if self.etiquetage == COORD:
                rows.append(
                    {
                        "nom": etud[0][0],
                        "prenom": etud[0][1],
                        "colonne": etud[1][0],
                        "ligne": etud[1][1],
                    }
                )
            else:
                rows.append({"nom": etud[0][0], "prenom": etud[0][1], "place": etud[1]})
        tab = GenTable(
            titles=titles,
            columns_ids=columns_ids,
            rows=rows,
            filename=filename,
            origin="Généré par %s le " % sco_version.SCONAME
            + scu.timedate_human_repr()
            + "",
            pdf_title=pdf_title,
            # pdf_shorttitle = '',
            preferences=sco_preferences.SemPreferences(
                self.moduleimpl_data["formsemestre_id"]
            ),
        )
        return tab.make_page(format="pdf", with_html_headers=False)

    def _one_header(self, worksheet):
        cells = [
            worksheet.make_cell("Nom", self.styles["2bi"]),
            worksheet.make_cell("Prénom", self.styles["2bi"]),
        ]
        if self.etiquetage == COORD:
            cells.append(worksheet.make_cell("Colonne", self.styles["2bi"]))
            cells.append(worksheet.make_cell("Ligne", self.styles["2bi"]))
        else:
            cells.append(worksheet.make_cell("Place", self.styles["2bi"]))
        return cells

    def _headers(self, worksheet, nb_listes):
        cells = []
        for _ in range(nb_listes):
            cells += self._one_header(worksheet)
            cells.append(worksheet.make_cell(""))
        worksheet.append_row(cells)

    def _make_styles(self, ws0, ws1):
        # polices
        font0 = Font(name="Calibri", bold=True, size=12)
        font1b = copy(font0)
        font1b.size = 9
        font1i = Font(name="Arial", italic=True, size=10)
        font1o = Font(name="Arial", outline=True, size=10)
        font2bi = Font(name="Arial", bold=True, italic=True, size=8)
        font2 = Font(name="Arial", size=10)

        # bordures
        side_double = Side(border_style="double", color=COLORS.BLACK.value)
        side_thin = Side(border_style="thin", color=COLORS.BLACK.value)

        # bordures
        border1t = Border(left=side_double, top=side_double, right=side_double)
        border1bb = Border(left=side_double, bottom=side_double, right=side_double)
        border1bm = Border(left=side_double, right=side_double)
        border1m = Border(left=side_double, bottom=side_thin, right=side_double)
        border2m = Border(top=side_thin, bottom=side_thin)
        border2r = Border(top=side_thin, bottom=side_thin, right=side_thin)
        border2l = Border(left=side_thin, top=side_thin, bottom=side_thin)
        border2b = Border(
            left=side_thin, top=side_thin, bottom=side_thin, right=side_thin
        )

        # alignements
        align_center_center = Alignment(horizontal="center", vertical="center")
        align_right_bottom = Alignment(horizontal="right", vertical="bottom")
        align_left_center = Alignment(horizontal="left", vertical="center")
        align_right_center = Alignment(horizontal="right", vertical="center")

        # patterns
        pattern = PatternFill(
            fill_type="solid", fgColor=sco_excel.COLORS.LIGHT_YELLOW.value
        )

        # styles
        self.styles = {
            "titres": sco_excel.excel_make_style(font_name="Arial", bold=True, size=12),
            "1t": ws0.excel_make_composite_style(
                font=font0, alignment=align_center_center, border=border1t
            ),
            "1m": ws0.excel_make_composite_style(
                font=font1b, alignment=align_center_center, border=border1m
            ),
            "1bm": ws0.excel_make_composite_style(
                font=font1b, alignment=align_center_center, border=border1bm
            ),
            "1bb": ws0.excel_make_composite_style(
                font=font1o, alignment=align_right_bottom, border=border1bb
            ),
            "2b": ws1.excel_make_composite_style(
                font=font1i, alignment=align_center_center, border=border2b
            ),
            "2bi": ws1.excel_make_composite_style(
                font=font2bi,
                alignment=align_center_center,
                border=border2b,
                fill=pattern,
            ),
            "2l": ws1.excel_make_composite_style(
                font=font2, alignment=align_left_center, border=border2l
            ),
            "2m1": ws1.excel_make_composite_style(
                font=font2, alignment=align_left_center, border=border2m
            ),
            "2m2": ws1.excel_make_composite_style(
                font=font2, alignment=align_right_center, border=border2m
            ),
            "2r": ws1.excel_make_composite_style(
                font=font2, alignment=align_right_center, border=border2r
            ),
        }

    def _titres(self, worksheet):
        datetime = time.strftime("%d/%m/%Y a %Hh%M")
        worksheet.append_single_cell_row(
            "Feuille placement etudiants éditée le %s" % datetime, self.styles["titres"]
        )
        for line, desceval in enumerate(self.desceval):
            if line in [1, 4, 7]:
                worksheet.append_blank_row()
            worksheet.append_single_cell_row(desceval, self.styles["titres"])
        worksheet.append_single_cell_row(
            "Date : %(jour)s - Horaire : %(heure_debut)s à %(heure_fin)s"
            % self.eval_data,
            self.styles["titres"],
        )

    def _feuille0(self, ws0, space):
        self._titres(ws0)
        # entetes colonnes - feuille0
        cells = [ws0.make_cell()]
        for col in range(self.nb_rangs):
            cells.append(ws0.make_cell("colonne %s" % (col + 1), self.styles["2b"]))
        ws0.append_row(cells)

        # etudiants - feuille0
        place = 1
        col = 0
        rang = 1
        # Chaque rang est affiché sur 3 lignes xlsx (notées A, B, C)
        # ligne A: le nom, ligne B: le prénom, ligne C: un espace ou la place
        cells_a = [ws0.make_cell(rang, self.styles["2b"])]
        cells_b = [ws0.make_cell("", self.styles["2b"])]
        cells_c = [ws0.make_cell("", self.styles["2b"])]
        row = 13  # première ligne de signature
        rang += 1
        for linetud in self.plan:
            cells_a.append(ws0.make_cell(linetud[0][0], self.styles["1t"]))  # nom
            cells_b.append(ws0.make_cell(linetud[0][1], self.styles["1m"]))  # prenom
            if self.etiquetage == COORD:
                cell_c = ws0.make_cell("", self.styles["1bb"])
            else:
                cell_c = ws0.make_cell("place %s" % place, self.styles["1bb"])
                place = place + 1
            cells_c.append(cell_c)
            ws0.set_row_dimension_height(row, space / 25)
            row += 3
            col += 1
            if col == self.nb_rangs:  # On a fini la rangée courante
                ws0.append_row(cells_a)  # on affiche les 3 lignes construites
                ws0.append_row(cells_b)
                ws0.append_row(cells_c)
                cells_a = [
                    ws0.make_cell(rang, self.styles["2b"])
                ]  # on réinitialise les 3 lignes
                cells_b = [ws0.make_cell("", self.styles["2b"])]
                cells_c = [ws0.make_cell("", self.styles["2b"])]
                col = 0
                rang += 1
            # publication du rang final incomplet
        ws0.append_row(cells_a)  # Affiche des 3 lignes (dernières lignes incomplètes)
        ws0.append_row(cells_b)
        ws0.append_row(cells_c)
        ws0.set_row_dimension_height(row, space / 25)

    def _feuille1(self, worksheet, maxlines):
        # etudiants - feuille1
        # structuration:
        # 1 page = maxlistes listes
        #   1 liste = 3 ou 4 colonnes(excel) (selon numbering) et (maximum maxlines) lignes
        maxlistes = 2  # nombre de listes par page
        # computes excel columns widths
        if self.etiquetage == COORD:
            gabarit = [16, 18, 6, 6, 2]
        else:
            gabarit = [16, 18, 12, 2]
        widths = []
        for _ in range(maxlistes):
            widths += gabarit
        worksheet.set_column_dimension_width(value=widths)
        nb_etu_restant = len(self.listetud)
        self._titres(worksheet)
        nb_listes = min(
            maxlistes, nb_etu_restant // maxlines + 1
        )  # nombre de colonnes dans la page
        self._headers(worksheet, nb_listes)
        # construction liste alphabétique
        # Affichage
        lines = [[] for _ in range(maxlines)]
        lineno = 0
        col = 0
        for etud in sorted(self.plan, key=lambda e: e[0][0]):  # tri alphabétique
            # check for skip of list or page
            if col > 0:  # add a empty cell between lists
                lines[lineno].append(worksheet.make_cell())
            lines[lineno].append(worksheet.make_cell(etud[0][0], self.styles["2l"]))
            lines[lineno].append(worksheet.make_cell(etud[0][1], self.styles["2m1"]))
            if self.etiquetage == COORD:
                lines[lineno].append(
                    worksheet.make_cell(etud[1][1], self.styles["2m2"])
                )
                lines[lineno].append(worksheet.make_cell(etud[1][0], self.styles["2r"]))
            else:
                lines[lineno].append(worksheet.make_cell(etud[1], self.styles["2r"]))
            lineno = lineno + 1
            if lineno >= maxlines:  # fin de liste
                col = col + 1
                lineno = 0
                if col >= maxlistes:  # fin de page
                    for line_cells in lines:
                        worksheet.append_row(line_cells)
                    lines = [[] for _ in range(maxlines)]
                    col = 0
                    worksheet.append_blank_row()
                    nb_etu_restant -= maxlistes * maxlines
                    nb_listes = min(
                        maxlistes, nb_etu_restant // maxlines + 1
                    )  # nombre de colonnes dans la page
                    self._headers(worksheet, nb_listes)
        for line_cells in lines:
            worksheet.append_row(line_cells)

    def _excel_feuille_placement(self):
        """Genere feuille excel pour placement des etudiants.
        E: evaluation (dict)
        lines: liste de tuples
                   (etudid, nom, prenom, etat, groupe, val, explanation)
        """
        sem_preferences = sco_preferences.SemPreferences()
        space = sem_preferences.get("feuille_placement_emargement")
        maxlines = sem_preferences.get("feuille_placement_positions")
        nb_rangs = int(self.nb_rangs)
        column_width_ratio = (
            1 / 250
        )  # changement d unités entre pyExcelerator et openpyxl

        workbook = ScoExcelBook()

        sheet_name_0 = "Emargement"
        ws0 = workbook.create_sheet(sheet_name_0)
        # ajuste largeurs colonnes (unite inconnue, empirique)
        width = 4500 * column_width_ratio
        if nb_rangs > 5:
            width = 22500 * column_width_ratio // nb_rangs

        ws0.set_column_dimension_width("A", 750 * column_width_ratio)
        for col in range(nb_rangs):
            ws0.set_column_dimension_width(
                "ABCDEFGHIJKLMNOPQRSTUVWXYZ"[col + 1 : col + 2], width
            )

        sheet_name_1 = "Positions"
        ws1 = workbook.create_sheet(sheet_name_1)

        self._make_styles(ws0, ws1)
        self._feuille0(ws0, space)
        self._feuille1(ws1, maxlines)
        return workbook.generate()
